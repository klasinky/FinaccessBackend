from openpyxl import load_workbook, Workbook
from rest_framework import viewsets, mixins, status
from rest_framework.generics import get_object_or_404
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
from api.permissions import IsMonthOwner
from core.models import Month, Category
from io import BytesIO


class AmountBaseCreateView(mixins.CreateModelMixin,
                           viewsets.GenericViewSet):
    permission_classes = [IsAuthenticated, IsMonthOwner, ]

    def dispatch(self, request, *args, **kwargs):
        id_month = kwargs.pop('id')
        self.month = get_object_or_404(Month, id=id_month)
        return super(AmountBaseCreateView, self).dispatch(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        try:
            category = request.data['category']
        except Exception as e:
            return Response({"category": ["Este campo es requerido."]},
                            status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        category = get_object_or_404(Category, id=category)

        self.perform_create(serializer, category=category)

        headers = self.get_success_headers(serializer.data)

        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer, category):
        serializer.save(category=category, month=self.month)


class AmountBaseUploadXLS(APIView):
    parser_classes = (FileUploadParser,)

    permission_classes = [IsAuthenticated, IsMonthOwner, ]
    model = None

    def dispatch(self, request, *args, **kwargs):
        id_month = kwargs.pop('id')
        self.month = get_object_or_404(Month, id=id_month)
        return super(AmountBaseUploadXLS, self).dispatch(request, *args, **kwargs)

    def post(self, request, format=None, *args, **kwargs):
        print("Entra en el post")
        try:
            file_obj = request.data["file"]
            #file_obj = request.FILES["file"]
            wb = load_workbook(filename=BytesIO(file_obj.read()))
        except Exception as e:
            print(e)
            return Response({"file": ["Formato incorrecto."]},
                            status=status.HTTP_400_BAD_REQUEST)

        sheet = wb.active
        data = {
            'inserts': [],
            'errors': []
        }
        for row_cells in sheet.iter_rows(min_row=2,
                                         min_col=0,
                                         max_col=4):
            if row_cells[0] is not None:
                name = row_cells[0].value
                description = row_cells[1].value
                amount = row_cells[2].value
                category_id = row_cells[3].value

                if name and description and amount and category_id:
                    try:
                        category = Category.objects.get(pk=int(category_id))
                        amount = float(amount)
                        model = self.model.objects.create(
                            name=name,
                            description=description,
                            amount=amount,
                            category=category,
                            month=self.month
                        )
                        model.save()
                        data['inserts'].append(
                            f"{name} - {description} - {amount} - {category.name}"
                        )
                    except ValueError as e:
                        print(e)
                        data['errors'].append({
                            'row': row_cells[2].coordinate,
                            'error': 'El amount tiene un formato erróneo'
                        })

                    except Category.DoesNotExist as e:
                        print(e)
                        data['errors'].append({
                            'row': row_cells[3].coordinate,
                            'error': 'La categoria no existe'
                        })
        return Response(data, status=status.HTTP_200_OK)


class AmountBaseDownloadXLS(APIView):
    permission_classes = [IsAuthenticated, IsMonthOwner, ]
    filename = None
    model = None

    def dispatch(self, request, *args, **kwargs):
        id_month = kwargs.pop('id')
        self.month = get_object_or_404(Month, id=id_month)
        return super(AmountBaseDownloadXLS, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        amount_list = self.model.objects.filter(month=self.month)
        amount_title_fields = ['name', 'description', 'amount', 'category']

        if amount_list:
            book = Workbook()
            sheet = book.active
            sheet.title = 'Sheet 1'
            sheet.append((amount_title_fields[0],
                          amount_title_fields[1],
                          amount_title_fields[2],
                          amount_title_fields[3]))
            for item in amount_list:
                sheet.append((item.name,
                              item.description,
                              f"{item.amount:.2f}{request.user.currency.symbol}", item.category.name))

            file_name = f"{self.filename}{self.month.date.year}-{self.month.date.month}.xlsx"
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            content = "attachment; filename=%s" % (file_name)
            response['Content-Disposition'] = content
            row = sheet[1]
            self.color_row(row)

            book.active = self.auto_size_xls(sheet)
            book.save(response)

            return response
        return Response({'detail':'No hay registros para descargar'},
                        status=status.HTTP_404_NOT_FOUND)

    def auto_size_xls(self, sheet):
        """
               Función encontrada en
               https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
            """
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value + 3
        return sheet

    def color_row(self, row):
        for cell in row:
            if cell.value:
                cell.font = Font(color="FFFFFF", size=12, bold=True)
                cell.fill = PatternFill("solid", fgColor="022B85")
                cell.alignment = Alignment(horizontal='center')
        return row
